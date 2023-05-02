#!/usr/bin/python3
"""
RestFul API bulk_load for users
"""
from flask import jsonify, request
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from api.views import app_views
from models.user import User
import os


UPLOAD_FOLDER = os.path.abspath("./uploads/")
ALLOWED_EXTENSIONS = set(["xlsx"])


def allowed_file(filname):
    """
    Method to validate the allowed extensions
    """
    if "." in filname and filname.split(".")[1].lower() in ALLOWED_EXTENSIONS:
        return True
    return False


@app_views.route("/bulkload", methods=["POST"])
def bulk_load():
    """
    Endpoint that loads the excel file to then carry out
    the bulk load of user data
    """
    if request.method == "POST":
        print("request.files {}".format(request.files))
        f = request.files["file"]
        if f.filename == "":
            return "No file selected."
        if f and allowed_file(f.filename):
            filename = secure_filename(f.filename)
            f.save(os.path.join(UPLOAD_FOLDER, filename))
            get_file = os.scandir(UPLOAD_FOLDER)
            with get_file as files_folder:
                for fichero in files_folder:
                    if fichero.name == 'carga_masiva_user.xlsx':
                        path = UPLOAD_FOLDER+'/'+fichero.name
                        workbook = load_workbook(path)
                        sheet = workbook.active
                        print(sheet)
                        for r in range(2, sheet.max_row + 1):
                            values = {
                                    'first_name': sheet.cell(r, 1).value,
                                    'id_document_type': sheet.cell(r, 2).value,
                                    'last_name': sheet.cell(r, 3).value,
                                    'email': sheet.cell(r, 4).value,
                                    'password': sheet.cell(r, 5).value,
                                    'number_document': sheet.cell(r, 6).value,
                                    'phone': sheet.cell(r, 7).value,
                                    'birth_date': sheet.cell(r, 8).value
                                    }
                            print(values)
                            instance = User(**values)
                            instance.new()
                        os.remove(path)
        else:
            return jsonify({'mensaje': 'File not allowed'}), 406
    return jsonify(instance.to_dict()), 201
